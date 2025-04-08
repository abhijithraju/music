from flask import Flask, request, render_template, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

EXCEL_FILE = 'comments.xlsx'

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Comment", "Timestamp"])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json
    name = data.get('name')
    comment = data.get('comment')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, comment, timestamp])
    wb.save(EXCEL_FILE)

    return jsonify({'status': 'success', 'timestamp': timestamp})

@app.route('/comments')
def get_comments():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    comments = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        comments.append({
            'name': row[0],
            'comment': row[1],
            'timestamp': row[2]
        })

    comments.reverse()
    return jsonify(comments)

if __name__ == '__main__':
    app.run(debug=True)
